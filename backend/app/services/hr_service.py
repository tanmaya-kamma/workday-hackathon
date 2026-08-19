"""
HR service — organization-wide queries,
reporting, statistics, leave policies,
and regional holiday calendars.

Uses Motor (PyMongo async) directly
for database operations on LMS database.
"""

from typing import (
    Optional,
    Dict,
    Any,
    List,
)

from io import BytesIO
from datetime import datetime

# pyrefly: ignore [missing-import]
from bson import ObjectId

from fastapi import (
    HTTPException,
    status,
    UploadFile,
)

from app.core.database import get_db
from app.core.security import hash_password

from app.schemas.user import UserProfile
from app.schemas.leave import (
    LeaveResponse,
    LeaveListResponse,
)
from app.schemas.hr import (
    HRDashboardStats,
    LeaveTypeDistribution,
    CreateEmployeeRequest,
)

from app.services.auth_service import (
    _doc_to_profile,
)

from app.services.leave_service import (
    _doc_to_response,
)

# ============================================================
# DYNAMIC ACCRUAL ENGINE
# ============================================================

from app.services.accrual_service import (
    AccrualService,
)


# ============================================================
# EMPLOYEE MANAGEMENT
# ============================================================


async def create_employee(
    data: CreateEmployeeRequest,
) -> UserProfile:
    """
    Create a new employee/manager record
    in LMS.users.

    Immediately after creating the employee,
    the dynamic accrual engine calculates and
    persists the employee's leave balance.

    Leave entitlement is NOT manually assigned
    from the frontend.

    The accrual engine is responsible for:

    - Date of joining
    - Organization leave policy
    - Accrual rules
    - Employee tenure
    - First-year proration
    - Regional calendar
    - Existing leave usage
    - Pending leave
    """

    db = get_db()

    # ========================================================
    # CLEAN EMAIL
    # ========================================================

    email_clean = (
        data.email.strip().lower()
    )

    # ========================================================
    # CHECK EXISTING EMAIL
    # ========================================================

    existing = await db.users.find_one(
        {
            "email": email_clean
        }
    )

    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"An employee with email "
                f"'{email_clean}' already exists."
            ),
        )

    # ========================================================
    # EMPLOYEE ID
    # ========================================================

    emp_id = data.employee_id

    if not emp_id or not emp_id.strip():

        count = await db.users.count_documents(
            {
                "role": data.role
            }
        )

        prefix = (
            "MGR"
            if data.role == "manager"
            else (
                "HR"
                if data.role == "hr"
                else "EMP"
            )
        )

        emp_id = (
            f"{prefix}{count + 1:03d}"
        )

    else:

        emp_id = (
            emp_id.strip().upper()
        )

        existing_emp = (
            await db.users.find_one(
                {
                    "employee_id": emp_id
                }
            )
        )

        if existing_emp:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    f"An employee with ID "
                    f"'{emp_id}' already exists."
                ),
            )

    # ========================================================
    # MANAGER OBJECT ID
    # ========================================================

    mgr_oid = None

    if (
        data.manager_id
        and data.manager_id.strip()
    ):

        try:
            mgr_oid = ObjectId(
                data.manager_id.strip()
            )

        except Exception:
            mgr_oid = None

    # ========================================================
    # PASSWORD
    # ========================================================

    pwd_raw = (
        data.password
        or "password123"
    )

    pwd_hash = hash_password(
        pwd_raw
    )

    # ========================================================
    # DATE OF JOINING
    # ========================================================

    doj = (
        data.date_of_joining
        or datetime.utcnow()
    )

    # ========================================================
    # USER DOCUMENT
    # ========================================================

    user_doc = {
        "employee_id": emp_id,

        "email": email_clean,

        "password_hash": pwd_hash,

        "full_name":
            data.full_name.strip(),

        "role":
            data.role,

        "department":
            data.department.strip(),

        "region":
            data.region.strip(),

        "manager_id":
            mgr_oid,

        "date_of_joining":
            doj,

        "is_active":
            True,
    }

    # ========================================================
    # INSERT USER FIRST
    # ========================================================

    result = await db.users.insert_one(
        user_doc
    )

    new_user_id = (
        result.inserted_id
    )

    user_doc["_id"] = (
        new_user_id
    )

    # ========================================================
    # DYNAMIC INITIAL LEAVE BALANCE
    # ========================================================

    """
    IMPORTANT:

    Do NOT manually insert:

        annual = 20
        sick = 12
        casual = 6

    anymore.

    The AccrualService is the source of truth.

    It calculates the employee's actual
    entitlement based on the employee's
    joining date and organization policy.
    """

    try:

        calculation_date = (
            datetime.utcnow().date()
        )

        # ----------------------------------------------------
        # CALCULATE
        # ----------------------------------------------------

        calculation = (
            AccrualService.calculate_employee_balance(
                employee_id=emp_id,
                as_of_date=calculation_date,
            )
        )

        # ----------------------------------------------------
        # SAVE
        # ----------------------------------------------------

        AccrualService.save_balance(
            employee=user_doc,
            calculation=calculation,
            year=calculation_date.year,
        )

        print(
            "[HR] Initial leave balance "
            "successfully calculated and "
            f"saved for {emp_id}."
        )

        print(
            "[HR] Calculation:",
            calculation,
        )

    except Exception as exc:

        # ----------------------------------------------------
        # IMPORTANT ERROR LOGGING
        # ----------------------------------------------------

        print(
            "[HR] ERROR: Employee was created "
            "but initial accrual calculation "
            f"failed for {emp_id}."
        )

        print(
            "[HR] Accrual error:",
            repr(exc),
        )

        # ----------------------------------------------------
        # CLEAN UP INVALID EMPLOYEE
        # ----------------------------------------------------

        """
        We don't want an employee existing
        without a corresponding leave balance.

        Therefore, if the mandatory initial
        balance calculation fails, remove
        the employee again and return an error.

        This makes employee creation atomic
        from the application's perspective.
        """

        try:

            await db.users.delete_one(
                {
                    "_id":
                        new_user_id
                }
            )

        except Exception as cleanup_exc:

            print(
                "[HR] WARNING: Could not "
                "rollback employee after "
                f"accrual failure: {cleanup_exc}"
            )

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=(
                "Employee could not be created "
                "because the leave accrual "
                "calculation failed. "
                f"Accrual error: {str(exc)}"
            ),
        )

    # ========================================================
    # RETURN CREATED EMPLOYEE
    # ========================================================

    return await _doc_to_profile(
        user_doc
    )


# ============================================================
# EMPLOYEE DIRECTORY
# ============================================================


async def get_all_employees() -> List[UserProfile]:
    """
    Retrieve all active users across
    the organization.

    User profiles should expose the latest
    leave balance calculated by the backend.
    """

    db = get_db()

    cursor = (
        db.users
        .find(
            {
                "is_active": True
            }
        )
        .sort(
            "full_name",
            1,
        )
    )

    employees = await cursor.to_list(
        length=1000
    )

    return [
        await _doc_to_profile(emp)
        for emp in employees
    ]


async def get_all_managers() -> List[UserProfile]:
    """
    Retrieve all active managers.
    """

    db = get_db()

    cursor = (
        db.users
        .find(
            {
                "role": "manager",
                "is_active": True,
            }
        )
        .sort(
            "full_name",
            1,
        )
    )

    managers = await cursor.to_list(
        length=1000
    )

    return [
        await _doc_to_profile(mgr)
        for mgr in managers
    ]


# ============================================================
# ORGANIZATIONAL LEAVES
# ============================================================


async def get_organizational_leaves(
    employee_id: Optional[str] = None,
    manager_id: Optional[str] = None,
    status: Optional[str] = None,
    leave_type: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
) -> LeaveListResponse:
    """
    Retrieve all leave requests with
    filtering and pagination.
    """

    db = get_db()

    query: Dict[str, Any] = {}

    # ========================================================
    # EMPLOYEE FILTER
    # ========================================================

    if employee_id:

        try:

            query["$or"] = [
                {
                    "employee_id":
                        employee_id
                },
                {
                    "employee_id":
                        ObjectId(
                            employee_id
                        )
                },
            ]

        except Exception:

            query[
                "employee_id"
            ] = employee_id

    # ========================================================
    # MANAGER FILTER
    # ========================================================

    if manager_id:

        try:

            query["$or"] = [
                {
                    "manager_id":
                        manager_id
                },
                {
                    "manager_id":
                        ObjectId(
                            manager_id
                        )
                },
            ]

        except Exception:

            query[
                "manager_id"
            ] = manager_id

    # ========================================================
    # OTHER FILTERS
    # ========================================================

    if status:
        query[
            "status"
        ] = status

    if leave_type:
        query[
            "leave_type"
        ] = leave_type

    # ========================================================
    # COUNT
    # ========================================================

    total = (
        await db.leave_requests.count_documents(
            query
        )
    )

    # ========================================================
    # PAGINATION
    # ========================================================

    skip = (
        (page - 1)
        * limit
    )

    cursor = (
        db.leave_requests
        .find(query)
        .sort(
            "created_at",
            -1,
        )
        .skip(skip)
        .limit(limit)
    )

    leaves = await cursor.to_list(
        length=limit
    )

    from app.services.leave_service import (
        _enrich_leaves_with_employee_names
    )

    items = (
        await _enrich_leaves_with_employee_names(
            leaves,
            db,
        )
    )

    return LeaveListResponse(
        items=items,
        total=total,
    )


# ============================================================
# HR STATISTICS
# ============================================================


async def get_leave_statistics() -> HRDashboardStats:
    """
    Compute aggregate leave statistics
    for the HR dashboard.
    """

    db = get_db()

    total_employees = (
        await db.users.count_documents(
            {
                "role": "employee",
                "is_active": True,
            }
        )
    )

    total_managers = (
        await db.users.count_documents(
            {
                "role": "manager",
                "is_active": True,
            }
        )
    )

    total_requests = (
        await db.leave_requests.count_documents(
            {}
        )
    )

    pending_requests = (
        await db.leave_requests.count_documents(
            {
                "status": "pending"
            }
        )
    )

    approved_requests = (
        await db.leave_requests.count_documents(
            {
                "status": "approved"
            }
        )
    )

    rejected_requests = (
        await db.leave_requests.count_documents(
            {
                "status": "rejected"
            }
        )
    )

    # ========================================================
    # LEAVE DISTRIBUTION
    # ========================================================

    annual_count = (
        await db.leave_requests.count_documents(
            {
                "leave_type": {
                    "$in": [
                        "annual",
                        "vacation",
                    ]
                }
            }
        )
    )

    sick_count = (
        await db.leave_requests.count_documents(
            {
                "leave_type":
                    "sick"
            }
        )
    )

    casual_count = (
        await db.leave_requests.count_documents(
            {
                "leave_type": {
                    "$in": [
                        "casual",
                        "personal",
                    ]
                }
            }
        )
    )

    unpaid_count = (
        await db.leave_requests.count_documents(
            {
                "leave_type":
                    "unpaid"
            }
        )
    )

    distribution = (
        LeaveTypeDistribution(
            annual=annual_count,
            sick=sick_count,
            casual=casual_count,
            unpaid=unpaid_count,
        )
    )

    return HRDashboardStats(
        total_employees=
            total_employees,

        total_managers=
            total_managers,

        total_requests=
            total_requests,

        pending_requests=
            pending_requests,

        approved_requests=
            approved_requests,

        rejected_requests=
            rejected_requests,

        leave_type_distribution=
            distribution,
    )


# ============================================================
# LEAVE POLICY CONFIGURATION
# ============================================================


async def get_leave_policies() -> Dict[str, Any]:
    """
    Get the organization's current
    leave policy configuration.

    If no policy exists yet, create
    a default configuration.
    """

    db = get_db()

    current_year = datetime.utcnow().year

    # The ACTIVE policy: either the legacy year-less doc or the
    # newest doc whose effective_year has already arrived. Docs
    # with a future effective_year are returned separately below.
    policy = (
        await db.leave_policies.find_one(
            {
                "organization_id": "default",
                "$or": [
                    {
                        "effective_year": {
                            "$exists": False
                        }
                    },
                    {
                        "effective_year": {
                            "$lte": current_year
                        }
                    },
                ],
            },
            sort=[("effective_year", -1)],
        )
    )

    # ========================================================
    # CREATE DEFAULT POLICY
    # ========================================================

    if not policy:

        policy = {
            "organization_id":
                "default",

            "region":
                "India",

            "annual_leave":
                20,

            "sick_leave":
                12,

            "casual_leave":
                6,

            "manager_approval_days":
                2,

            "hr_direct_approval_days":
                6,

            "updated_at":
                datetime.utcnow(),
        }

        await db.leave_policies.insert_one(
            policy
        )

    # ========================================================
    # REMOVE MONGO ID
    # ========================================================

    policy.pop(
        "_id",
        None,
    )

    # ========================================================
    # SERIALIZE DATE
    # ========================================================

    if policy.get(
        "updated_at"
    ):

        policy[
            "updated_at"
        ] = (
            policy[
                "updated_at"
            ].isoformat()
        )

    # ========================================================
    # UPCOMING (FUTURE-YEAR) POLICIES
    # ========================================================

    upcoming_cursor = db.leave_policies.find(
        {
            "organization_id": "default",
            "effective_year": {
                "$gt": current_year
            },
        }
    ).sort("effective_year", 1)

    upcoming = []

    async for doc in upcoming_cursor:

        doc.pop("_id", None)

        if isinstance(doc.get("updated_at"), datetime):
            doc["updated_at"] = doc["updated_at"].isoformat()

        upcoming.append(doc)

    policy["upcoming_policies"] = upcoming

    return policy


async def update_leave_policies(
    data: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Update organization-wide
    leave policy configuration.
    """

    db = get_db()

    current_year = datetime.utcnow().year

    effective_year = int(data["effective_year"])

    # ========================================================
    # CURRENT-YEAR POLICIES ARE IMMUTABLE
    # ========================================================
    #
    # The policy in effect for the current year already governs
    # live balances, validations, and approvals. It cannot be
    # replaced mid-year; new policies apply from a future year.

    if effective_year <= current_year:

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"The {current_year} leave policy is already in "
                f"effect and cannot be added or changed — employee "
                f"balances are being calculated from it right now. "
                f"New policies can only be added for "
                f"{current_year + 1} and onward."
            ),
        )

    update_data = {
        "effective_year": effective_year,

        "annual_leave":
            float(
                data[
                    "annual_leave"
                ]
            ),

        "sick_leave":
            float(
                data[
                    "sick_leave"
                ]
            ),

        "casual_leave":
            float(
                data[
                    "casual_leave"
                ]
            ),

        "manager_approval_days":
            int(
                data[
                    "manager_approval_days"
                ]
            ),

        "hr_direct_approval_days":
            int(
                data[
                    "hr_direct_approval_days"
                ]
            ),

        "updated_at":
            datetime.utcnow(),
    }

    await db.leave_policies.update_one(
        {
            "organization_id":
                "default",

            "effective_year":
                effective_year,
        },

        {
            "$set":
                update_data,

            "$setOnInsert": {
                "organization_id":
                    "default",

                "region":
                    "India",
            },
        },

        upsert=True,
    )

    # ========================================================
    # VERSION THE ACCRUAL POLICY DOCUMENTS
    # ========================================================
    #
    # AccrualService selects policies by effective_from/effective_to,
    # so a future policy is stored as a new dated version: the
    # currently open version is closed on Dec 31 of the prior year
    # and the new version opens on Jan 1 of the effective year.
    # Current-year balances are untouched; the engine switches
    # over automatically at the year boundary.

    new_start = datetime(effective_year, 1, 1)

    prior_end = datetime(effective_year - 1, 12, 31)

    accrual_sync = [
        ("VACATION", float(data["annual_leave"])),
        ("SICK", float(data["sick_leave"])),
        ("PERSONAL", float(data["casual_leave"])),
    ]

    for leave_type, entitlement in accrual_sync:

        # Latest version that starts before the new policy —
        # the template for config we don't collect on this form.
        template = await db.policies.find_one(
            {
                "leave_type": leave_type,
                "effective_from": {
                    "$lt": new_start
                },
            },
            sort=[("effective_from", -1)],
        )

        if template is None:
            continue

        # Close the open-ended predecessor at the year boundary.
        if template.get("effective_to") is None:

            await db.policies.update_one(
                {
                    "_id": template["_id"]
                },
                {
                    "$set": {
                        "effective_to": prior_end
                    }
                },
            )

        balance = dict(
            template.get("balance") or {}
        )

        maximum = balance.get("maximum")

        if maximum is not None and maximum < entitlement:
            balance["maximum"] = entitlement

        future_version = {
            "policy_id": f"{leave_type}_{effective_year}",
            "leave_type": leave_type,
            "annual_entitlement": entitlement,
            "tenure_rules": [],
            "accrual": template.get("accrual"),
            "proration": template.get("proration"),
            "rounding": template.get("rounding"),
            "carry_forward": template.get("carry_forward"),
            "balance": balance,
            "effective_from": new_start,
            "effective_to": None,
        }

        await db.policies.update_one(
            {
                "leave_type": leave_type,
                "effective_from": new_start,
            },
            {
                "$set": future_version
            },
            upsert=True,
        )

    return await get_leave_policies()


# ============================================================
# REGIONAL HOLIDAY CALENDAR
# ============================================================


async def upload_regional_calendar(
    region: str,
    file: UploadFile,
) -> Dict[str, Any]:
    """
    Read an Excel regional holiday
    calendar and store the parsed
    holidays in MongoDB.

    Expected Excel columns:

        Date
        Holiday Name
    """

    # ========================================================
    # OPENPYXL
    # ========================================================

    try:

        from openpyxl import (
            load_workbook
        )

    except ImportError:

        raise HTTPException(
            status_code=500,
            detail=(
                "openpyxl is not installed. "
                "Run: pip install openpyxl"
            ),
        )

    # ========================================================
    # READ FILE
    # ========================================================

    try:

        contents = (
            await file.read()
        )

        workbook = load_workbook(
            filename=BytesIO(
                contents
            ),
            data_only=True,
        )

        worksheet = (
            workbook.active
        )

        rows = list(
            worksheet.iter_rows(
                values_only=True
            )
        )

        if not rows:

            raise HTTPException(
                status_code=400,
                detail=(
                    "The Excel file is empty."
                ),
            )

        # ====================================================
        # READ HEADERS
        # ====================================================

        headers = [
            (
                str(value)
                .strip()
                .lower()
                if value is not None
                else ""
            )
            for value in rows[0]
        ]

        date_index = None
        name_index = None

        # ====================================================
        # FIND COLUMNS
        # ====================================================

        for index, header in enumerate(
            headers
        ):

            if header in {
                "date",
                "holiday date",
                "holiday_date",
            }:

                date_index = index

            if header in {
                "holiday",
                "holiday name",
                "holiday_name",
                "name",
            }:

                name_index = index

        # ====================================================
        # VALIDATE COLUMNS
        # ====================================================

        if date_index is None:

            raise HTTPException(
                status_code=400,
                detail=(
                    "Excel must contain "
                    "a 'Date' column."
                ),
            )

        if name_index is None:

            raise HTTPException(
                status_code=400,
                detail=(
                    "Excel must contain "
                    "a 'Holiday Name' column."
                ),
            )

        # ====================================================
        # PARSE HOLIDAYS
        # ====================================================

        holidays = []

        for row in rows[1:]:

            if not row:
                continue

            holiday_date = (
                row[date_index]
                if date_index <
                len(row)
                else None
            )

            holiday_name = (
                row[name_index]
                if name_index <
                len(row)
                else None
            )

            if (
                not holiday_date
                or not holiday_name
            ):
                continue

            # ------------------------------------------------
            # EXCEL DATETIME
            # ------------------------------------------------

            if hasattr(
                holiday_date,
                "date",
            ):

                holiday_date = (
                    holiday_date.date()
                )

            # ------------------------------------------------
            # PYTHON DATE
            # ------------------------------------------------

            if hasattr(
                holiday_date,
                "isoformat",
            ):

                holiday_date = (
                    holiday_date.isoformat()
                )

            else:

                holiday_date = str(
                    holiday_date
                )

            holidays.append(
                {
                    "date":
                        holiday_date,

                    "name":
                        str(
                            holiday_name
                        ).strip(),
                }
            )

        # ====================================================
        # VALIDATE DATA
        # ====================================================

        if not holidays:

            raise HTTPException(
                status_code=400,
                detail=(
                    "No valid holiday records "
                    "were found in the Excel file."
                ),
            )

        # ====================================================
        # DETERMINE YEAR
        # ====================================================

        try:

            year = int(
                holidays[0][
                    "date"
                ][:4]
            )

        except Exception:

            year = (
                datetime.utcnow().year
            )

        # ====================================================
        # SAVE TO MONGODB
        # ====================================================

        db = get_db()

        calendar_doc = {
            "region":
                region,

            "year":
                year,

            "filename":
                file.filename,

            "holidays":
                holidays,

            "holiday_count":
                len(holidays),

            "updated_at":
                datetime.utcnow(),
        }

        await db.regional_calendars.update_one(
            {
                "region":
                    region,

                "year":
                    year,
            },

            {
                "$set":
                    calendar_doc
            },

            upsert=True,
        )

        # ====================================================
        # RESPONSE
        # ====================================================

        return {
            "success":
                True,

            "message":
                (
                    f"{len(holidays)} holidays "
                    "uploaded successfully."
                ),

            "region":
                region,

            "year":
                year,

            "holiday_count":
                len(holidays),

            "filename":
                file.filename,
        }

    except HTTPException:
        raise

    except Exception as exc:

        raise HTTPException(
            status_code=400,
            detail=(
                "Unable to process Excel "
                f"file: {exc}"
            ),
        )